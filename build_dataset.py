"""Build TRENDS.xlsx from the 30 raw MOPH 506 line-list files.

Inputs:
    SRC/Dengue/Dengue YYYY.xlsx                10 yearly files
    SRC/Chikungunya/Chikungunya YYYY.xlsx      10 yearly files
    SRC/Hand Foot and Mouth/HFM YYYY.xlsx      10 yearly files
    --master                                   province reference + populations

Output:
    --out  TRENDS.xlsx with three sheets: weekly data, province data,
           data dictionary.

Pipeline stages:
    eda      one-line summary per source file
    stage 1  normalise schemas, concatenate, filter onset dates to window
    stage 2  match province names to the official P-code
    stage 3  assign each case to an epidemiological week and aggregate
    final    join populations, compute rates, write the workbook
"""
import argparse
import os
import sys
import pandas as pd
import numpy as np



# moph 506 disease codes used by both the legacy (2016-2024) and the
# 2025 files; the 2025 files add an ICD-10 column but preserve these codes
DISEASE_CODE = {
    66: 'Dengue fever (DF)',
    26: 'Dengue haemorrhagic fever (DHF)',
    27: 'Dengue shock syndrome (DSS)',
    71: 'HFMD',
    84: 'Chikungunya',
}

# folders inside SRC and the file-name prefix used for each disease
DISEASE_FOLDERS = [
    ('Dengue',      'Dengue',              'Dengue'),
    ('Chikungunya', 'Chikungunya',         'Chikungunya'),
    ('HFMD',        'Hand Foot and Mouth', 'HFM'),
]

# column-name maps from each source schema onto a canonical name
LEGACY_TO_CANON = {
    'DISEASE':    'disease_code',
    'SEX':        'sex',
    'YEAR':       'age_y',
    'MONTH':      'age_m',
    'DAY':        'age_d',
    'MARIETAL':   'marital_status',
    'RACE':       'ethnicity',
    'occ_name':   'occupation',
    'Pro_Name':   'province_raw',
    'AMP_NAME':   'district',
    'TAM_NAME':   'subdistrict',
    'METROPOL':   'setting',
    'TYPE':       'encounter_type',
    'RESULT':     'outcome',
    'DATESICK':   'onset_date',
    'DATEDEFINE': 'report_date',
    'DATEDEATH':  'death_date',
}
MODERN_TO_CANON = {
    'epidem_report_group_code': 'disease_code',
    'gender':                   'sex',
    'age_y':                    'age_y',
    'age_m':                    'age_m',
    'age_d':                    'age_d',
    'marital_status_name':      'marital_status',
    'nationality':              'ethnicity',
    'occupation':               'occupation',
    'Pro_Name':                 'province_raw',
    'AMP_NAME':                 'district',
    'TUM_NAME':                 'subdistrict',
    'municipal_name':           'setting',
    'patient_type':             'encounter_type',
    'epidem_person_status_name':'outcome',
    'onset_date':               'onset_date',
    'report_datetime':          'report_date',
    'death_date':               'death_date',
}

# moph source files use Thai script for province names (Pro_Name), so
# the primary join is Thai -> rtsd Thai. the lookup below handles
# pre-romanised input as a fallback only.
NAME_FIXES = {
    'Bangkok Metropolis': 'Bangkok',
    'Buriram':            'Buri Ram',
    'Chainat':            'Chai Nat',
    'Chonburi':           'Chon Buri',
    'Lopburi':            'Lop Buri',
    'Nong Bua Lamphu':    'Nong Bua Lam Phu',
    'Phang Nga':          'Phangnga',
    'Prachinburi':        'Prachin Buri',
    'Sisaket':            'Si Sa Ket',
}



# eda
# return a list of (disease_label, year, path) tuples for the 30 files
def list_source_files(src):
    out = []
    for label, folder, prefix in DISEASE_FOLDERS:
        for y in range(2016, 2026):
            p = os.path.join(src, folder, f'{prefix} {y}.xlsx')
            out.append((label, y, p))
    return out


# one-line summary per file: rows, schema, missing onset dates, year span, distinct disease codes


def eda_pass(files):
    rows = []
    for label, year, path in files:
        if not os.path.exists(path):
            rows.append((label, year, 'MISSING', '', '', '', ''))
            continue
        d = pd.read_excel(path)
        schema = 'legacy' if 'DISEASE' in d.columns else 'modern'
        onset = 'DATESICK' if schema == 'legacy' else 'onset_date'
        n_missing = int(d[onset].isna().sum())
        d_codes = sorted(d[('DISEASE' if schema == 'legacy'
                             else 'epidem_report_group_code')].dropna().unique())
        codes_str = ','.join(str(int(c)) for c in d_codes if not pd.isna(c))
        # year span derived from onset dates that parse
        ds = pd.to_datetime(d[onset], errors='coerce')
        ys = ds.dt.year.dropna()
        yspan = f'{int(ys.min())}-{int(ys.max())}' if len(ys) else ''
        rows.append((label, year, schema, len(d), n_missing, codes_str, yspan))

    print('=' * 84)
    print('EDA pass - one row per source file')
    print('=' * 84)
    print(f'{"Disease":<13}{"File year":>10}{"Schema":>9}{"Rows":>10}'
          f'{"Missing onset":>15}{"Codes":>13}{"Onset yr range":>17}')
    for r in rows:
        print(f'{r[0]:<13}{r[1]:>10}{r[2]:>9}'
              f'{(r[3] if r[3]!="" else 0):>10,}'
              f'{(r[4] if r[4]!="" else 0):>15,}{r[5]:>13}{r[6]:>17}')
    print()



# stage 1: schema normalisation, concatenation, date filtering
# read one source file and return a dataframe with the canonical column names regardless of legacy or modernised schema
def normalise_one(path):
    d = pd.read_excel(path)
    if 'DISEASE' in d.columns:
        d = d.rename(columns=LEGACY_TO_CANON)[list(LEGACY_TO_CANON.values())]
    else:
        d = d.rename(columns=MODERN_TO_CANON)[list(MODERN_TO_CANON.values())]
    # parse dates
    for c in ('onset_date', 'report_date', 'death_date'):
        d[c] = pd.to_datetime(d[c], errors='coerce')
    # disease code as integer (drop rows where it cannot be coerced)
    d['disease_code'] = pd.to_numeric(d['disease_code'], errors='coerce')
    return d


# load all files, return one combined dataframe with canonical schema


def stage1_consolidate(files):
    print('Stage 1 - consolidating and cleaning')
    pieces = []
    for label, year, path in files:
        d = normalise_one(path)
        d['source_file_year'] = year
        d['disease_group'] = label
        pieces.append(d)
    df = pd.concat(pieces, ignore_index=True)
    print(f'  raw rows across 30 files: {len(df):,}')

    # drop rows with missing onset date or out-of-window onset
    keep = (df['onset_date'].dt.year >= 2016) & (df['onset_date'].dt.year <= 2025)
    n_drop_missing = int(df['onset_date'].isna().sum())
    n_drop_window  = int((~keep & df['onset_date'].notna()).sum())
    df = df[keep].copy()
    print(f'  dropped (missing onset date): {n_drop_missing:,}')
    print(f'  dropped (onset outside 2016 to 2025 window): {n_drop_window:,}')
    print(f'  retained for aggregation: {len(df):,}')
    print()
    return df



# stage 2: province name reconciliation, p-code attachment
# match each row to the official p-code
def stage2_geography(df, rtsd):
    print('Stage 2 - geographic harmonisation')
    df = df.copy()

    thai_lookup = (rtsd[['P-code', 'Province name (Thai)',
                          'Province name (English)']]
                    .rename(columns={'Province name (Thai)':    'province_raw',
                                     'Province name (English)': 'Province'}))
    df = df.merge(thai_lookup, on='province_raw', how='left')
    n_thai = int(df['P-code'].notna().sum())
    print(f'  rows matched via Thai-script join: {n_thai:,}')

    # romanised english fallback for any rows still unmatched and not blank
    unmatched_mask = df['P-code'].isna() & df['province_raw'].notna()
    if unmatched_mask.any():
        fix = df.loc[unmatched_mask, 'province_raw'].astype(str).replace(NAME_FIXES)
        eng_lookup = (rtsd[['P-code', 'Province name (English)']]
                       .rename(columns={'Province name (English)': 'Province'}))
        recovered = (pd.DataFrame({'Province': fix.values},
                                    index=fix.index)
                     .merge(eng_lookup, on='Province', how='left'))
        df.loc[unmatched_mask, 'P-code']   = recovered['P-code'].values
        df.loc[unmatched_mask, 'Province'] = recovered['Province'].values
        n_recovered = int(unmatched_mask.sum() - df.loc[unmatched_mask, 'P-code'].isna().sum())
        print(f'  rows recovered via romanised-name fallback '
              f'(nine-name fix applied where needed): {n_recovered:,}')

    blank = int(df['province_raw'].isna().sum())
    still_missing = int(df['P-code'].isna().sum())
    df = df[df['P-code'].notna()].copy()
    print(f'  rows dropped with blank province field: {blank:,}')
    if still_missing - blank > 0:
        print(f'  rows dropped with unmatched non-blank province: '
              f'{still_missing - blank:,}')
    print(f'  retained after geography join: {len(df):,}')
    print()
    return df



# stage 3: epidemiological week, aggregation to province-week
# assign each row to the canonical CDC MMWR (Year, Epi week) bucket by matching its onset_date against the (Week start, Week end) intervals taken from the master clean-data sheet
def assign_epi_week(d, week_skeleton):
    skel = (week_skeleton[['Year','Epi week','Week start','Week end']]
             .drop_duplicates()
             .copy())
    skel['Week start'] = pd.to_datetime(skel['Week start'])
    skel['Week end']   = pd.to_datetime(skel['Week end'])
    skel = skel.sort_values('Week start').reset_index(drop=True)
    d = d.copy()
    d['onset_date'] = pd.to_datetime(d['onset_date'])
    d = d.sort_values('onset_date').reset_index(drop=True)
    merged = pd.merge_asof(d, skel[['Week start','Year','Epi week','Week end']],
                            left_on='onset_date', right_on='Week start',
                            direction='backward')
    # drop rows whose onset_date lies past the matched Week end (defensive)
    keep = merged['onset_date'] <= merged['Week end']
    n_dropped = int((~keep).sum())
    if n_dropped:
        print(f'  rows beyond the last week skeleton interval: '
              f'{n_dropped:,} dropped')
    return merged[keep].copy()


# assign canonical (Year, Epi week), count cases per province-week per disease label, then pivot to one column per disease


def stage3_aggregate(df, week_skeleton):
    print('Stage 3 - temporal aggregation to province-week')
    df = assign_epi_week(df, week_skeleton)
    df['Disease label'] = df['disease_code'].map(DISEASE_CODE)

    counts = (df.groupby(['Year','Epi week','Week start','Week end',
                          'P-code','Province','Disease label'])
                .size().rename('cases').reset_index())
    wide = counts.pivot_table(
        index=['Year','Epi week','Week start','Week end','P-code','Province'],
        columns='Disease label', values='cases', fill_value=0).reset_index()
    wide.columns.name = None

    for col in ['Dengue fever (DF)', 'Dengue haemorrhagic fever (DHF)',
                'Dengue shock syndrome (DSS)', 'Chikungunya', 'HFMD']:
        if col not in wide.columns: wide[col] = 0

    wide['Dengue total'] = (wide['Dengue fever (DF)']
                                  + wide['Dengue haemorrhagic fever (DHF)']
                                  + wide['Dengue shock syndrome (DSS)'])
    print(f'  rows in long aggregate: {len(counts):,}')
    print(f'  rows in wide aggregate (province-week): {len(wide):,}')
    print()
    return wide



# final stage: complete grid, population join, rate computation, workbook
# 77 provinces x 527 epi weeks = 40,579 cells
def build_complete_grid(clean):
    weeks_df = (clean[['Year','Epi week','Week start','Week end']]
                  .drop_duplicates()
                  .copy())
    weeks_df['Week start'] = pd.to_datetime(weeks_df['Week start'])
    weeks_df['Week end']   = pd.to_datetime(weeks_df['Week end'])
    weeks_df = weeks_df.sort_values(['Year','Epi week']).reset_index(drop=True)
    prov = (clean[['P-code','Province (English)']]
              .drop_duplicates()
              .rename(columns={'Province (English)':'Province'})
              .sort_values('P-code')
              .reset_index(drop=True))
    grid = weeks_df.assign(_key=1).merge(prov.assign(_key=1),
                                          on='_key').drop(columns='_key')
    return grid


# join year-specific population at the (P-code, Year) level


def attach_population(weekly, pop_long):
    weekly = weekly.merge(pop_long, on=['P-code','Year'], how='left')
    for d, c in [('Dengue total',  'Dengue rate per 100,000'),
                 ('Chikungunya',         'Chikungunya rate per 100,000'),
                 ('HFMD',                'HFMD rate per 100,000')]:
        weekly[c] = (weekly[d] / weekly['Population'] * 1e5).round(3)
    return weekly


# write the three-sheet TRENDS.xlsx with the canonical intro line on row 1 and the data starting on row 3


def write_workbook(weekly, province, dictionary, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook(); wb.remove(wb.active)
    HEADER_FILL = PatternFill('solid', fgColor='305496')
    HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    def add_sheet(name, df, intro):
        ws = wb.create_sheet(title=name)
        c = ws.cell(row=1, column=1, value=intro)
        c.font = Font(name='Calibri', size=11, bold=True)
        c.alignment = Alignment(wrap_text=True, vertical='center')
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=max(1, len(df.columns)))
        ws.row_dimensions[1].height = 60
        for j, col in enumerate(df.columns, start=1):
            cc = ws.cell(row=2, column=j, value=str(col))
            cc.fill = HEADER_FILL; cc.font = HEADER_FONT
            cc.alignment = Alignment(horizontal='left', vertical='center')
        import datetime as _dt
        date_cols = {j for j, c in enumerate(df.columns, start=1)
                     if str(c) in ('Week start', 'Week end')}
        for i, row in enumerate(df.itertuples(index=False), start=3):
            for j, val in enumerate(row, start=1):
                if j in date_cols and isinstance(val, _dt.datetime):
                    val = val.date()
                cell = ws.cell(row=i, column=j, value=val)
                if j in date_cols and val is not None:
                    cell.number_format = 'yyyy-mm-dd'
        for j, col in enumerate(df.columns, start=1):
            sample = df.iloc[:50, j-1].astype(str).fillna('')
            width = min(max([len(str(col))] + [len(s) for s in sample]) + 2, 60)
            ws.column_dimensions[get_column_letter(j)].width = width
        ws.freeze_panes = 'A3'

    add_sheet('weekly data', weekly, intro=(
        'Analysis-ready weekly notifiable disease counts and pre-computed '
        'incidence rates per 100,000 for Thailand, 2016 to 2025. Disease '
        'counts: MOPH 506. Year-specific province populations used for the '
        'rate columns are provided in the province data sheet (join on P-code).'))
    add_sheet('province data', province, intro=(
        'Royal Thai Survey Department ADM-1 polygons for the 77 Thai '
        'provinces (boundaries valid from 22 January 2022, version v01), '
        'distributed via the United Nations OCHA Common Operational Dataset, '
        'paired with year-specific civil-registration population for each '
        'province from the Department of Provincial Administration.'))
    add_sheet('data dictionary', dictionary, intro=(
        'Variable definitions for the weekly data and province data sheets.'))
    wb.save(out_path)


# build the canonical 22-row data dictionary dataframe


def data_dictionary():
    weekly_rows = [
        ('Year','integer','year','Calendar year, 2016 to 2025.'),
        ('Month','integer','month','Calendar month, 1 to 12.'),
        ('Epi week','integer','week',
         'Epidemiological week within the year, 1 to 52 or 53.'),
        ('Week start','date','YYYY-MM-DD',
         'First calendar date of the epidemiological week.'),
        ('Week end','date','YYYY-MM-DD',
         'Last calendar date of the epidemiological week.'),
        ('Health region','integer','region (1 to 13)',
         'Ministry of Public Health health region grouping the province.'),
        ('Dengue fever (DF)','integer','cases',
         'Reported dengue fever cases (ICD-10 A90).'),
        ('Dengue haemorrhagic fever (DHF)','integer','cases',
         'Reported dengue haemorrhagic fever cases (ICD-10 A91).'),
        ('Dengue shock syndrome (DSS)','integer','cases',
         'Reported dengue shock syndrome cases (ICD-10 A91.1).'),
        ('Dengue total','integer','cases',
         'Total dengue cases, sum of DF, DHF, and DSS for the same '
         'province-week.'),
        ('Chikungunya','integer','cases',
         'Reported chikungunya virus disease cases (ICD-10 A92.0).'),
        ('HFMD','integer','cases',
         'Reported hand, foot, and mouth disease cases (ICD-10 B08.4, B08.5).'),
        ('Dengue rate per 100,000','numeric','rate per 100,000',
         'Weekly dengue rate per 100,000 population (computed from the '
         'Dengue total case count).'),
        ('Chikungunya rate per 100,000','numeric','rate per 100,000',
         'Weekly chikungunya virus disease rate per 100,000 population.'),
        ('HFMD rate per 100,000','numeric','rate per 100,000',
         'Weekly hand, foot, and mouth disease rate per 100,000 population.'),
    ]
    shared_rows = [
        ('Country','string','',
         'Country to which the (province, week) observation belongs. All rows '
         'in the Thailand release are "Thailand". The Country column anchors '
         'per-country subsetting and enables future expansion of TRENDS to '
         'additional Southeast Asian countries.'),
        ('P-code','string','TH10 to TH96',
         'Official Thai administrative code for the province (Royal Thai '
         'Survey Department). Join key between the two sheets.'),
        ('Province','string','name',
         'Province name in English (Royal Thai Survey Department spelling).'),
    ]
    province_rows = [
        ('Centroid latitude','numeric','degrees',
         'WGS84 (EPSG:4326) latitude of the province polygon centroid.'),
        ('Centroid longitude','numeric','degrees',
         'WGS84 (EPSG:4326) longitude of the province polygon centroid.'),
        ('Area (km²)','numeric','km²',
         'Province area in square kilometres, computed in EPSG:3035.'),
        ('Population 2016 to Population 2025','integer','persons',
         'Year-specific civil-registration population for the province, one '
         'column per year, from the Department of Provincial Administration. '
         'Join on P-code to recover the matching denominator for any '
         '(province, year) row in weekly data.'),
    ]
    rows = []
    for v,t,u,d in weekly_rows:   rows.append([v,'weekly data',t,u,d])
    for v,t,u,d in shared_rows:   rows.append([v,'weekly data, province data',t,u,d])
    for v,t,u,d in province_rows: rows.append([v,'province data',t,u,d])
    return pd.DataFrame(rows,
        columns=['variable','sheet','type','unit','definition'])



# main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--src', required=True,
                    help='folder holding Dengue/, Chikungunya/, Hand Foot '
                         'and Mouth/ subfolders with the 30 MOPH files')
    ap.add_argument('--master', required=True,
                    help='Disease Data (2016 - 2025) master workbook used '
                         'for RTSD ADM-1 reference and yearly DOPA populations')
    ap.add_argument('--out', required=True,
                    help='destination path for TRENDS.xlsx')
    ap.add_argument('--country', default='Thailand',
                    help='country label to write into the Country column '
                         '(default: Thailand). Use the canonical short name '
                         'in English when building releases for other '
                         'Southeast Asian countries.')
    args = ap.parse_args()

    files = list_source_files(args.src)
    missing = [p for _, _, p in files if not os.path.exists(p)]
    if missing:
        print(f'ERROR: missing {len(missing)} source file(s):', file=sys.stderr)
        for p in missing: print(f'  {p}', file=sys.stderr)
        sys.exit(1)

    # run an EDA pass first to summarise every input file
    eda_pass(files)

    # load the province reference table and the yearly populations
    rtsd = pd.read_excel(args.master, sheet_name='RTSD ADM-1', skiprows=2)
    clean = pd.read_excel(args.master, sheet_name='Clean Data', header=2)
    pop_long = (clean[['Province (English)','Year','Population']]
                .drop_duplicates()
                .rename(columns={'Province (English)':'Province'})
                .merge(rtsd[['P-code','Province name (English)']]
                          .rename(columns={'Province name (English)':'Province'}),
                       on='Province', how='inner')
                [['P-code','Year','Population']])

    # stages 1 to 3 build the long-form aggregate
    df = stage1_consolidate(files)
    df = stage2_geography(df, rtsd)
    weekly_long = stage3_aggregate(df, week_skeleton=clean)

    # expand to the complete grid, then attach populations
    grid = build_complete_grid(clean)
    weekly = grid.merge(weekly_long,
                         on=['Year','Epi week','Week start','Week end',
                             'P-code','Province'], how='left').fillna(0)
    weekly = attach_population(weekly, pop_long)

    # attach month and health region columns from the master clean-data sheet
    hr = (clean[['Province (English)','Year','Epi week','Month','Health region']]
            .drop_duplicates()
            .rename(columns={'Province (English)':'Province'}))
    weekly = weekly.merge(hr, on=['Province','Year','Epi week'], how='left')

    # prepend the Country column; default value comes from --country
    weekly['Country'] = args.country
    weekly_order = [
        'Country',
        'Year','Month','Epi week','Week start','Week end','Health region',
        'P-code','Province',
        'Dengue fever (DF)','Dengue haemorrhagic fever (DHF)',
        'Dengue shock syndrome (DSS)','Dengue total',
        'Chikungunya','HFMD',
        'Dengue rate per 100,000','Chikungunya rate per 100,000',
        'HFMD rate per 100,000',
    ]
    for c in weekly_order:
        if c not in weekly.columns: weekly[c] = pd.NA
    weekly = weekly[weekly_order].astype({
        'Year':'Int64','Month':'Int64','Epi week':'Int64','Health region':'Int64',
        'Dengue fever (DF)':'Int64','Dengue haemorrhagic fever (DHF)':'Int64',
        'Dengue shock syndrome (DSS)':'Int64','Dengue total':'Int64',
        'Chikungunya':'Int64','HFMD':'Int64',
    })

    # build the province data sheet
    pop_wide = (pop_long.pivot(index='P-code', columns='Year',
                                values='Population').reset_index())
    pop_wide.columns = ['P-code'] + [f'Population {int(y)}'
                                       for y in pop_wide.columns[1:]]
    prov = rtsd[['P-code','Province name (English)',
                  'Centroid latitude','Centroid longitude','Area (km²)']
                ].rename(columns={'Province name (English)':'Province'})
    prov = prov.merge(pop_wide, on='P-code', how='left')
    # prepend the Country column on the province sheet too
    prov.insert(0, 'Country', args.country)

    # build the data dictionary
    dd = data_dictionary()

    print('Final dataset')
    print(f'  weekly data:      {weekly.shape}')
    print(f'  province data:    {prov.shape}')
    print(f'  data dictionary:  {dd.shape}')
    print(f'  total dengue: {int(weekly["Dengue total"].sum()):,}')
    print(f'  total chikungunya:      {int(weekly["Chikungunya"].sum()):,}')
    print(f'  total HFMD:             {int(weekly["HFMD"].sum()):,}')
    print()

    write_workbook(weekly, prov, dd, args.out)
    print(f'Wrote: {args.out}')


if __name__ == '__main__':
    main()